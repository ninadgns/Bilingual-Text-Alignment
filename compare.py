"""Align an English/Traditional-Chinese HKEX PDF pair entry-by-entry and write to Excel.

Usage:
    python3 compare.py <en_folder> <zh_folder> [out_folder]

    Matches PDFs by filename across the two folders and writes one aligned Excel
    file per pair into out_folder (default: ./output).

Strategy:
- Extract words from each page with their (x, y) positions; group into visual lines.
- Group lines into "entries" using a combined rule:
    new entry starts when (vertical gap > ENTRY_GAP) AND (line starts with a quote or
    a numbered-rule token like "1.02A").
  Sub-notes ("Note:") and lettered sub-items ("(a)") fail the pattern check; inline
  quoted aliases like “AFRC” fail the gap check. Both stay merged into their parent.
- Align EN and ZH entries by position. Sanity check each row using two independent signals:
    1. Glossary anchor: the English term appears in parens inside the ZH entry
       (e.g. ZH `“申請版本” (Application Proof)` ↔ EN `“Application Proof”`).
    2. Rule-ID identity: both texts start with the same numbered rule (e.g. `4.21`).
"""
import argparse
import os
import re
from dataclasses import dataclass
from typing import Optional

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# Per-PDF column split (x-coordinate). Words with x0 < split = left col (term),
# >= split = right col (definition). Holds across HKEX Main Board rules PDFs.
EN_COL_SPLIT = 180
ZH_COL_SPLIT = 145

# Page regions to ignore (header decorations and page footers)
HEADER_Y_MAX = 140   # chapter title decoration on page 1 (multi-line titles can reach y~131)
FOOTER_Y_MIN_OFFSET = 50  # exclude bottom 50px (page numbers like "1 – 1")

# Right-margin decoration filter: some chapters have vertical sidebar text (e.g. "第/三/章"
# running down the right edge near x≈441 on a 465-wide page) that otherwise merges into
# mid-page content lines and suppresses entry boundaries.
SIDEBAR_MARGIN = 30  # drop words whose x0 > page.width - SIDEBAR_MARGIN

# Horizontal y-tolerance for grouping words into the same visual line
LINE_Y_TOL = 3

# Smart open-quote characters that mark the start of a defined term.
# `《` is needed for ZH definitions like `《歐盟國際財務匯報準則》(EU-IFRS)`.
# Mid-paragraph `《...》` references are filtered by the gap-threshold check.
OPEN_QUOTE_CHARS = ("“", "《")

# Numbered-rule pattern (e.g. 1.01, 1.02A) — must be the FIRST word in the left column.
# Uses \b instead of $ so it still matches when the rule number is glued to an opening
# quote with no space (ch2's `2.02A“`, ch3's `3.09G《`).
NUMBERED_RULE_RE = re.compile(r"^\d+\.\d+[A-Z]?\b")

# Captures the rule-ID at the start of an entry's text, for the rule-identity sanity check.
RULE_ID_RE = re.compile(r"^\s*(\d+\.\d+[A-Z]?)\b")

# Vertical-gap threshold (px). A new entry boundary requires BOTH a gap above this AND a
# matching start-pattern (quote / rule number). This prevents within-entry quoted aliases
# like `“AFRC”` (small gap) from being split off as separate entries.
# Tuned: chapter 1 ZH p6 has between-entry gaps as small as 16.9; within-entry tops out ~13.
ENTRY_GAP = 14


@dataclass
class Entry:
    page: int
    text: str           # full reconstructed text of the entry
    term: str = ""      # for two-column entries: the left-column term
    body: str = ""      # for two-column entries: the right-column definition
    en_anchor: str = "" # for ZH entries: the English term found in parens (sanity-check key)


def group_words_into_lines(words):
    """Return list of (top, [words]) sorted by top, words grouped by similar top."""
    if not words:
        return []
    words = sorted(words, key=lambda w: (w["top"], w["x0"]))
    lines = []
    current_top = words[0]["top"]
    current = []
    for w in words:
        if abs(w["top"] - current_top) <= LINE_Y_TOL:
            current.append(w)
        else:
            lines.append((current_top, current))
            current_top = w["top"]
            current = [w]
    if current:
        lines.append((current_top, current))
    return lines


def split_line(line_words, col_split):
    """Split a line's words into (left_text, right_text) by x0 threshold."""
    line_words = sorted(line_words, key=lambda w: w["x0"])
    left = " ".join(w["text"] for w in line_words if w["x0"] < col_split)
    right = " ".join(w["text"] for w in line_words if w["x0"] >= col_split)
    return left, right


def has_entry_start_pattern(line_words, col_split: int) -> bool:
    """Return True if the line's leftmost word looks like a new-entry marker:
    starts with a smart open-quote in the left column, or matches a numbered-rule token."""
    if not line_words:
        return False
    leftmost = min(line_words, key=lambda w: w["x0"])
    if leftmost["x0"] >= col_split:
        return False
    text = leftmost["text"]
    if text.startswith(OPEN_QUOTE_CHARS):
        return True
    if NUMBERED_RULE_RE.match(text):
        return True
    return False


def extract_entries(path: str, col_split: int, lang: str) -> list[Entry]:
    """Boundary rule: a new entry starts when (vertical gap > ENTRY_GAP) AND
    (line starts with a quote/rule pattern). Sub-notes and lettered sub-items
    fail the pattern check; inline aliases like “AFRC” fail the gap check."""
    entries: list[Entry] = []
    current_lines: list[tuple[float, list]] = []
    current_page = 1
    prev_top: Optional[float] = None
    with pdfplumber.open(path) as pdf:
        for pi, page in enumerate(pdf.pages):
            page_num = pi + 1
            footer_y = page.height - FOOTER_Y_MIN_OFFSET
            header_y = HEADER_Y_MAX if page_num == 1 else 0
            sidebar_x = page.width - SIDEBAR_MARGIN
            words = [
                w for w in page.extract_words()
                if header_y < w["top"] < footer_y and w["x0"] < sidebar_x
            ]
            lines = group_words_into_lines(words)
            # Page break resets the gap signal — treat first line on a new page as gapped
            page_first = True
            for top, line_words in lines:
                gap = float("inf") if page_first or prev_top is None else (top - prev_top)
                page_first = False
                if current_lines and gap > ENTRY_GAP and has_entry_start_pattern(line_words, col_split):
                    entries.append(_build_entry(current_page, current_lines, col_split, lang))
                    current_lines = []
                if not current_lines:
                    current_page = page_num
                current_lines.append((top, line_words))
                prev_top = top
    if current_lines:
        entries.append(_build_entry(current_page, current_lines, col_split, lang))
    return entries


def _build_entry(page_num: int, entry_lines, col_split: int, lang: str) -> Entry:
    """Reconstruct an entry's text from its line groups."""
    left_parts: list[str] = []
    right_parts: list[str] = []
    full_parts: list[str] = []
    for _, line_words in entry_lines:
        left, right = split_line(line_words, col_split)
        if left:
            left_parts.append(left)
        if right:
            right_parts.append(right)
        if left and right:
            full_parts.append(f"{left}  {right}")
        else:
            full_parts.append(left or right)

    term = " ".join(left_parts).strip()
    body = " ".join(right_parts).strip()
    text = "\n".join(full_parts).strip()

    en_anchor = extract_en_anchor(term) if lang == "zh" else ""

    return Entry(page=page_num, text=text, term=term, body=body, en_anchor=en_anchor)


def extract_en_anchor(text: str) -> str:
    """Return the English term inside the first balanced parens that contains >=3 ASCII letters.
    Handles nested parens like `(Companies (Winding Up...) Ordinance)`."""
    i = text.find("(")
    while i >= 0:
        depth = 1
        j = i + 1
        while j < len(text) and depth > 0:
            if text[j] == "(":
                depth += 1
            elif text[j] == ")":
                depth -= 1
            j += 1
        if depth == 0:
            content = text[i + 1:j - 1]
            if sum(1 for c in content if c.isascii() and c.isalpha()) >= 3:
                return re.sub(r"\s+", " ", content).strip()
        i = text.find("(", i + 1)
    return ""


def en_term_of(entry: Entry) -> str:
    """Return the first quoted English term in the entry's left column (sanity-check key)."""
    m = re.search(r"[“”\"]([^“”\"]+)[“”\"]", entry.term)
    if m:
        return re.sub(r"\s+", " ", m.group(1)).strip()
    return ""


def rule_id_of(entry: Entry) -> str:
    """Return the leading numbered-rule identifier in the entry's text (e.g. '2.02A'), or ''."""
    m = RULE_ID_RE.match(entry.text)
    return m.group(1) if m else ""


_NORMALIZE_RE = re.compile(r"[\s\-–—_]+")


def anchors_match(en_term: str, zh_anchor: str) -> bool:
    """Lenient comparison: case-insensitive, ignores whitespace/dashes, accepts substring matches.
    Handles `Code on Share Buy-backs` vs `Code on Share Buy- backs or Share Buy- backs Code`."""
    a = _NORMALIZE_RE.sub("", en_term.lower())
    b = _NORMALIZE_RE.sub("", zh_anchor.lower())
    if not a or not b:
        return False
    return a in b or b in a


def sanity_label(en: Optional[Entry], zh: Optional[Entry]) -> tuple[str, str]:
    """Return (tag, category) where category is one of:
    'anchor-ok' | 'anchor-x' | 'rule-ok' | 'rule-x' | 'none'."""
    en_term = en_term_of(en) if en else ""
    zh_anchor = zh.en_anchor if zh else ""
    if en_term and zh_anchor:
        if anchors_match(en_term, zh_anchor):
            return "OK (anchor)", "anchor-ok"
        return f"X anchor en={en_term!r} zh={zh_anchor!r}", "anchor-x"

    en_rule = rule_id_of(en) if en else ""
    zh_rule = rule_id_of(zh) if zh else ""
    if en_rule and zh_rule:
        if en_rule == zh_rule:
            return f"OK ({en_rule})", "rule-ok"
        return f"X rule en={en_rule} zh={zh_rule}", "rule-x"

    return "-", "none"


def write_excel(rows, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Aligned"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    headers = ["Line", "Page (EN)", "Page (ZH)", "English", "Chinese", "Match?"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
    wrap = Alignment(wrap_text=True, vertical="top")
    for i, row in enumerate(rows, start=2):
        for col, val in enumerate(row, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.alignment = wrap
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 60
    ws.column_dimensions["E"].width = 60
    ws.column_dimensions["F"].width = 22
    ws.freeze_panes = "A2"
    wb.save(out_path)


def align(en_pdf: str, zh_pdf: str, out_xlsx: str) -> None:
    print(f"EN: {en_pdf}")
    print(f"ZH: {zh_pdf}")

    en_entries = extract_entries(en_pdf, EN_COL_SPLIT, lang="en")
    zh_entries = extract_entries(zh_pdf, ZH_COL_SPLIT, lang="zh")
    print(f"EN entries: {len(en_entries)}    ZH entries: {len(zh_entries)}")

    rows = []
    counts = {"anchor-ok": 0, "anchor-x": 0, "rule-ok": 0, "rule-x": 0, "none": 0}
    n = max(len(en_entries), len(zh_entries))
    for i in range(n):
        en = en_entries[i] if i < len(en_entries) else None
        zh = zh_entries[i] if i < len(zh_entries) else None
        tag, category = sanity_label(en, zh)
        counts[category] += 1
        rows.append([
            i + 1,
            en.page if en else "",
            zh.page if zh else "",
            en.text if en else "",
            zh.text if zh else "",
            tag,
        ])

    print(f"sanity: anchor OK={counts['anchor-ok']}  anchor X={counts['anchor-x']}  "
          f"rule OK={counts['rule-ok']}  rule X={counts['rule-x']}  no-signal={counts['none']}")
    write_excel(rows, out_xlsx)
    print(f"Wrote {out_xlsx} ({len(rows)} rows)")


def main():
    parser = argparse.ArgumentParser(description=__doc__.split("\n\n")[0])
    parser.add_argument("en_folder", help="Folder containing English PDFs")
    parser.add_argument("zh_folder", help="Folder containing Traditional-Chinese PDFs")
    parser.add_argument("out_folder", nargs="?", default="output",
                        help="Output folder for .xlsx files (default: ./output)")
    args = parser.parse_args()

    os.makedirs(args.out_folder, exist_ok=True)

    en_pdfs = sorted(
        f for f in os.listdir(args.en_folder) if f.lower().endswith(".pdf")
    )
    if not en_pdfs:
        print(f"No PDFs found in {args.en_folder}")
        return

    for filename in en_pdfs:
        en_path = os.path.join(args.en_folder, filename)
        zh_path = os.path.join(args.zh_folder, filename)
        if not os.path.exists(zh_path):
            print(f"SKIP {filename}: no matching ZH file at {zh_path}")
            continue
        stem = os.path.splitext(filename)[0].replace(" ", "_").lower()
        out_xlsx = os.path.join(args.out_folder, f"{stem}_aligned.xlsx")
        print(f"\n=== {filename} ===")
        align(en_path, zh_path, out_xlsx)


if __name__ == "__main__":
    main()
